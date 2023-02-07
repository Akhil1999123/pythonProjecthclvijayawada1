import openpyxl
wb=openpyxl.load_workbook("C:\pythonProjecthclvijayawada\dataset\Book1.xlsx")
#print(wb)
#sheet=wb['Sheet1']
#print(sheet)
#wb=opx.workbook
sheet=wb.active
data=sheet['A3'].value
sheet['A3'].value='hello'
print(data)
data3=sheet.cell(row=2,column=3).value
data1=sheet['A1:A10']

print(data)
print(data1)
print(data3)
#acess all cells in row
print(sheet.max_row)
print(sheet.min_row)

for i in range(2,8):
    print(sheet.cell(row=i,column=1).value)
