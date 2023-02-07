from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
#sheet.title="hcl"
#sheet["A1"].value=10
#sheet["B2"].value="hcl"
#sheet.cell(row=3,column=3).value="hcl data"
#wb.save('C:\pythonProjecthclvijayawada\dataset\Book1.xlsx')
j=0
for  i  in range(10,60,10):
    j=j+1
    sheet.cell(row=j,column=1).value=i
#j=0
#for row in sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=5):
  #  j=j+100
 #   for d in row:
  #      d.value=j
    #for r in row :
        #r.value=1
wb.save('C:\pythonProjecthclvijayawada\dataset\Book1.xlsx')